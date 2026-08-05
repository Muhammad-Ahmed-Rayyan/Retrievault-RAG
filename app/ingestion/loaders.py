from pathlib import Path
from typing import List

from langchain_core.documents import Document

SUPPORTED_EXTENSIONS = {".pdf", ".docx", ".txt", ".xlsx"}

class UnsupportedFileTypeError(Exception):
    """Raised when a file extension isn't in SUPPORTED_EXTENSIONS."""
    pass

from openpyxl import load_workbook
from langchain_core.documents import Document as LCDocument


class _ExcelRowLoader:
    """
    Lightweight XLSX loader using openpyxl directly, avoiding the
    heavier and slower 'unstructured' parsing stack. Converts each
    sheet into row-wise text so semantic search can still find values.
    """

    def __init__(self, file_path: str):
        self.file_path = file_path

    def load(self) -> list:
        workbook = load_workbook(self.file_path, data_only=True)
        documents = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            if not rows:
                continue

            headers = [str(h) if h is not None else "" for h in rows[0]]

            batch_size = 20
            for i in range(1, len(rows), batch_size):
                batch = rows[i:i + batch_size]
                lines = []
                for row in batch:
                    row_text = ", ".join(
                        f"{headers[j]}: {val}" for j, val in enumerate(row)
                        if j < len(headers) and val is not None
                    )
                    if row_text:
                        lines.append(row_text)

                if lines:
                    content = f"Sheet: {sheet_name}\n" + "\n".join(lines)
                    documents.append(LCDocument(
                        page_content=content,
                        metadata={"sheet_name": sheet_name, "row_start": i},
                    ))

        return documents


def load_document(file_path: str) -> List[Document]:
    """
    Load a single document from disk based on its file extension.

    Args:
        file_path: Absolute or relative path to the source file.

    Returns:
        A list of LangChain Document objects (one per page/section,
        depending on loader behavior).

    Raises:
        UnsupportedFileTypeError: If the extension isn't supported.
        FileNotFoundError: If the file doesn't exist.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    extension = path.suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise UnsupportedFileTypeError(
            f"'{extension}' is not supported. "
            f"Supported types: {', '.join(SUPPORTED_EXTENSIONS)}"
        )

    loader = _get_loader(str(path), extension)
    documents = loader.load()

    # Attach normalized source metadata for traceability during retrieval
    for doc in documents:
        doc.metadata["source_file"] = path.name
        doc.metadata["file_type"] = extension.lstrip(".")

    return documents


def _get_loader(file_path: str, extension: str):
    """Factory that maps a file extension to its LangChain loader."""
    loader_map = {
        ".pdf": lambda p: PyPDFLoader(p),
        ".docx": lambda p: Docx2txtLoader(p),
        ".txt": lambda p: TextLoader(p, encoding="utf-8"),
        ".xlsx": lambda p: _ExcelRowLoader(p),
    }
    return loader_map[extension](file_path)


def load_documents_from_directory(directory: str) -> List[Document]:
    """
    Load and combine all supported documents from a directory.

    Args:
        directory: Path to a directory containing source documents.

    Returns:
        Combined list of Document objects from every supported file found.
    """
    dir_path = Path(directory)
    all_documents: List[Document] = []
    skipped_files: List[str] = []

    for file_path in dir_path.iterdir():
        if file_path.is_file() and file_path.suffix.lower() in SUPPORTED_EXTENSIONS:
            try:
                all_documents.extend(load_document(str(file_path)))
            except Exception as e:
                skipped_files.append(f"{file_path.name}: {e}")

    if skipped_files:
        print(f"[loaders] Skipped {len(skipped_files)} file(s) due to errors:")
        for entry in skipped_files:
            print(f"  - {entry}")

    return all_documents